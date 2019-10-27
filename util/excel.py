# encoding=utf-8
from openpyxl import *
import os.path
from proj_var.var import ProjDirPath
from openpyxl.styles import NamedStyle, Font, colors
from calendar import *


class Excel:
    def __init__(self, excel_file_path):
        if os.path.exists(excel_file_path):
            self.excel_file_path = excel_file_path
            self.wb = load_workbook(self.excel_file_path)
        else:
            print("%s文件路径不存在，请重新设定" % excel_file_path)

    def set_sheet_by_name(self, sheet_name):
        if sheet_name in self.wb.sheetnames:
            self.sheet = self.wb[sheet_name]
        else:
            print("%s sheet不存在，请重新指定 sheet 名称进行设定" % sheet_name)

    def set_sheet_by_index(self, index):
        if isinstance(index, int) and 1 <= index <= len(self.get_all_sheet_names()):
            sheet_name = self.get_all_sheet_names()[index-1]
            self.sheet = self.wb[sheet_name]
        else:
            print("%s sheet 序号不存在，请重新指定 sheet 序号进行设定" % index)

    def get_current_sheet_name(self):
        return self.sheet.title

    def get_all_sheet_names(self):
        return self.wb.sheetnames

    def get_rows_object(self):
        return list(self.sheet.rows)

    def get_cols_object(self):
        return list(self.sheet.columns)

    def get_row(self, row_no):
        if isinstance(row_no, int) and 1 <= row_no <= len(self.get_rows_object()):
            return self.get_rows_object()[row_no-1]
        else:
            print("%s 行号不存在，请重新设定行号读取！" % row_no)

    def get_col(self, col_no):
        if isinstance(col_no, int) and 1 <= col_no <= len(self.get_cols_object()):
            return self.get_cols_object()[col_no - 1]
        else:
            print("%s 列号不存在，请重新设定行号读取！" % col_no)

    def get_cell_value(self, row_no, col_no):
        if isinstance(row_no, int) and isinstance(col_no, int) and \
            1 <= row_no <= len(self.get_rows_object()) and \
                1 <= row_no <= len(self.get_cols_object()):
            return self.sheet.cell(row=row_no, column=col_no).value
        else:
            print("%s,%s 行号或者列号不存在，请重新设定行号或者列表读取！" % (row_no, col_no))

    def write_cell_value(self, row_no, col_no, value, color=None):
        if isinstance(row_no, int) and isinstance(col_no, int):
            if color is None:
                font = Font(bold=False, size=10, color=colors.BLACK)
                self.sheet.cell(row=row_no, column=col_no).font = font
                self.sheet.cell(row=row_no, column=col_no).value = value
            elif color == "green":
                font = Font(bold=True, size=13, color=colors.GREEN)
                self.sheet.cell(row=row_no, column=col_no).font = font
                self.sheet.cell(row=row_no, column=col_no).value = value
            elif color == "red":
                font = Font(bold=True, size=13, color=colors.RED)
                self.sheet.cell(row=row_no, column=col_no).font = font
                self.sheet.cell(row=row_no, column=col_no).value = value
            self.wb.save(self.excel_file_path)
        else:
            print("%s,%s 行号或者列号不是数字，请重新设定行号或者列表读取！" % (row_no, col_no))

    def write_current_time(self, row_no, col_no):
        if isinstance(row_no, int) and isinstance(col_no, int):
            self.sheet.cell(row=row_no, column=col_no).value = get_current_date_and_time()
            self.wb.save(self.excel_file_path)

if __name__ == "__main__":
    excel_file_path = ProjDirPath+r"\test_data\testdata.xlsx"
    print(excel_file_path)
    excel_obj = Excel(excel_file_path)
    Excel(excel_file_path)
    excel_obj.set_sheet_by_name("首页")
    print(excel_obj.get_current_sheet_name())
    print(excel_obj.get_all_sheet_names())
    excel_obj.set_sheet_by_index(2)
    print(excel_obj.get_current_sheet_name())
    print(len(excel_obj.get_rows_object()))
    print(excel_obj.get_rows_object())
    print(len(excel_obj.get_cols_object()))
    print(excel_obj.get_row(3))
    print(excel_obj.get_col(2))
    print(excel_obj.get_cell_value(2,2))
    excel_obj.write_cell_value(8,12,"你好！")
    excel_obj.write_cell_value(9, 12, "你好","red")
    excel_obj.write_cell_value(10, 12, "你好","green")
    excel_obj.write_current_time(11,12)